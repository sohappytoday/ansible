import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BASE_DIR = Path("/app")
JSON_DIR = BASE_DIR / "json"
TODAY = datetime.now().strftime("%Y%m%d")
EXCEL_FILE = BASE_DIR / "server_spec" / f"server-spec-{TODAY}.xlsx"


def get_json_files():
    return sorted(JSON_DIR.glob("*.json"))


def load_rows():
    rows = []

    for json_file in get_json_files():
        with json_file.open("r", encoding="utf-8") as f:
            data = json.load(f)

        disk = data.get("disk", [])
        if isinstance(disk, list):
            disk_text = "\n".join(disk)
        else:
            disk_text = str(disk)

        rows.append({
            "Server": data.get("server", "N/A"),
            "Hostname": data.get("hostname", "N/A"),
            "Internal IP": data.get("internal_ip", "N/A"),
            "External IP": data.get("external_ip", "N/A"),
            "Gateway": data.get("gateway", "N/A"),
            "OS": data.get("os", "N/A"),
            "CPU Model": data.get("cpu_model_name", "N/A"),
            "Physical CPU Cores": data.get("cpu_physical_core_count", "N/A"),
            "CPU Thread/Core": data.get("cpu_thread_per_core", "N/A"),
            "RAM": data.get("ram", "N/A"),
            "Disk": disk_text,
        })

    return rows


def create_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Server Spec"

    headers = [
        "Server",
        "Hostname",
        "Internal IP",
        "External IP",
        "Gateway",
        "OS",
        "CPU Model",
        "Physical CPU Cores",
        "CPU Thread/Core",
        "RAM",
        "Disk",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([row.get(header, "N/A") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 18,
        "B": 22,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 30,
        "G": 50,
        "H": 16,
        "I": 18,
        "J": 12,
        "K": 55,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45

    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)


if __name__ == "__main__":
    rows = load_rows()

    if not rows:
        raise SystemExit("No JSON files found. Expected: output/json/*.json")

    create_excel(rows)
    print(f"Created: {EXCEL_FILE}")
